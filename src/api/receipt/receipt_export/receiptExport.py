# -*- coding: utf-8 -*-
# SPDX-License-Identifier: MIT
# Copyright (c) 2026 Home Kakeibo System Contributors

import io
import os
import uuid
import base64
from datetime import datetime
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ReceiptExportService:
    """レシート検索結果をExcelまたはPDFとして出力するサービス。"""

    def __init__(self):
        self.store = {}

    def prepare(self, request):
        """出力対象データを一時保存し、ダウンロードページURLを返す。"""
        export_type = self.text((request or {}).get("type")).lower()
        if export_type not in ("excel", "pdf"):
            return {"statusCode": 400, "body": {"errorMessage": "出力形式が不正です。"}}

        rows = self.normalize_rows((request or {}).get("rows") or [])
        if not rows:
            return {"statusCode": 400, "body": {"errorMessage": "出力対象のデータがありません。"}}

        token = uuid.uuid4().hex
        self.store[token] = {
            "type": export_type,
            "rows": rows,
            "createdAt": datetime.now(),
        }
        return {"statusCode": 200, "body": {"url": f"/export/receipt/page/{token}"}}

    def prepare_file(self, request):
        """Generate the requested file without relying on Lambda memory."""
        export_type = self.text((request or {}).get("type")).lower()
        if export_type not in ("excel", "pdf"):
            return {"statusCode": 400, "body": {"errorMessage": "出力形式が不正です。"}}

        rows = self.normalize_rows((request or {}).get("rows") or [])
        if not rows:
            return {"statusCode": 400, "body": {"errorMessage": "出力対象のデータがありません。"}}

        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        if export_type == "excel":
            content = self.build_excel(rows)
            filename = f"レシート検索結果_{now}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            content = self.build_pdf(rows)
            filename = f"レシート検索結果_{now}.pdf"
            media_type = "application/pdf"

        return {
            "statusCode": 200,
            "body": {
                "filename": filename,
                "mediaType": media_type,
                "contentBase64": base64.b64encode(content).decode("ascii"),
            },
        }

    def page_html(self, token):
        """ブラウザでダウンロードを開始するための中間HTMLを作成する。"""
        if token not in self.store:
            return None
        return f"""
<!doctype html>
<html lang="ja">
  <head>
    <meta charset="utf-8">
    <title>レシート出力</title>
    <style>
      body {{ margin: 0; min-height: 100vh; display: grid; place-items: center; font-family: "Yu Gothic", "Meiryo", sans-serif; background: #f8fafc; color: #111827; }}
      main {{ width: min(520px, calc(100vw - 40px)); padding: 32px; border: 1px solid #dbeafe; border-radius: 8px; background: #fff; box-shadow: 0 18px 45px rgba(15, 23, 42, .10); }}
      h1 {{ margin: 0 0 12px; font-size: 22px; }}
      p {{ margin: 0 0 20px; color: #475569; line-height: 1.7; }}
      a {{ color: #1d4ed8; font-weight: 700; }}
    </style>
  </head>
  <body>
    <main>
      <h1>ダウンロードを開始しています</h1>
      <p>ファイルの保存が始まらない場合は、下のリンクをクリックしてください。</p>
      <a id="downloadLink" href="/export/receipt/file/{token}">ファイルをダウンロード</a>
    </main>
    <script>
      window.addEventListener("load", () => document.getElementById("downloadLink").click());
    </script>
  </body>
</html>
"""

    def build_file(self, token):
        """保存済みトークンから出力ファイル本体とヘッダ情報を作る。"""
        export_data = self.store.get(token)
        if not export_data:
            return None

        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        if export_data["type"] == "excel":
            filename = f"レシート検索結果_{now}.xlsx"
            return {
                "content": self.build_excel(export_data["rows"]),
                "filename": filename,
                "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }

        filename = f"レシート検索結果_{now}.pdf"
        return {
            "content": self.build_pdf(export_data["rows"]),
            "filename": filename,
            "media_type": "application/pdf",
        }

    def attachment_headers(self, filename):
        """日本語ファイル名を含むダウンロードヘッダを作成する。"""
        fallback = "receipt_report" + os.path.splitext(filename)[1]
        return {
            "Content-Disposition": f"attachment; filename={fallback}; filename*=UTF-8''{quote(filename)}"
        }

    def normalize_rows(self, rows):
        """画面から渡された検索結果を出力用の安定した行形式へ変換する。"""
        normalized = []
        for row in rows or []:
            if not isinstance(row, dict):
                continue
            normalized.append({
                "receiptDate": self.text(row.get("receiptDate")),
                "receiptTime": self.text(row.get("receiptTime")),
                "invoiceRegistrationNumber": self.text(row.get("invoiceRegistrationNumber")),
                "supplierName": self.text(row.get("supplierName")),
                "category1": self.text(row.get("category1")),
                "category2": self.text(row.get("category2")),
                "itemName": self.text(row.get("itemName")),
                "quantity": self.number(row.get("quantity")),
                "unitPrice": self.number(row.get("unitPrice")),
                "taxRate": self.number(row.get("taxRate")),
                "totalPrice": self.number(row.get("totalPrice")),
            })
        return normalized

    def build_excel(self, rows):
        """明細、概要、分類集計、月別集計を含むExcelファイルを作成する。"""
        workbook = Workbook()
        detail = workbook.active
        detail.title = "明細"
        detail.append(["日付", "時刻", "登録番号", "店舗", "分類", "小分類", "商品名", "数量", "単価", "税率", "金額"])
        for row in rows:
            detail.append([
                row["receiptDate"],
                row["receiptTime"],
                row["invoiceRegistrationNumber"],
                row["supplierName"],
                row["category1"],
                row["category2"],
                row["itemName"],
                row["quantity"],
                row["unitPrice"],
                self.format_tax_rate(row["taxRate"]),
                row["totalPrice"],
            ])
        self.style_sheet(detail, [14, 10, 20, 24, 16, 18, 34, 10, 12, 10, 14])
        for cell in detail["I"][1:] + detail["K"][1:]:
            cell.number_format = '#,##0"円"'

        total, category_totals, month_totals = self.summary(rows)
        summary = workbook.create_sheet("概要")
        summary.append(["項目", "値"])
        summary.append(["明細件数", len(rows)])
        summary.append(["合計金額", total])
        summary.append(["分類数", len(category_totals)])
        summary.append(["出力日時", datetime.now().strftime("%Y/%m/%d %H:%M")])
        self.style_sheet(summary, [20, 24])
        summary["B3"].number_format = '#,##0"円"'

        by_category = workbook.create_sheet("分類集計")
        by_category.append(["分類", "金額"])
        for category, amount in sorted(category_totals.items(), key=lambda item: item[1], reverse=True):
            by_category.append([category, amount])
        self.style_sheet(by_category, [24, 16])
        for cell in by_category["B"][1:]:
            cell.number_format = '#,##0"円"'

        by_month = workbook.create_sheet("月別集計")
        by_month.append(["月", "金額"])
        for month, amount in sorted(month_totals.items()):
            by_month.append([month, amount])
        self.style_sheet(by_month, [16, 16])
        for cell in by_month["B"][1:]:
            cell.number_format = '#,##0"円"'

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def style_sheet(self, sheet, widths):
        """Excelシートのヘッダ、罫線、列幅を整える。"""
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(name="Yu Gothic", bold=True, color="FFFFFF")
        body_font = Font(name="Yu Gothic", color="1F2937")
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.font = body_font
                    if cell.row % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor="F8FAFC")

    def build_pdf(self, rows):
        """検索結果を見やすい複数ページのPDFレポートに変換する。"""
        total, category_totals, month_totals = self.summary(rows)
        pages = [self.pdf_cover_page(rows, total, category_totals, month_totals)]
        detail_rows = sorted(rows, key=lambda row: (row.get("receiptDate") or "", row.get("supplierName") or ""))
        offset = 0
        while offset < len(detail_rows):
            page_rows = []
            used_height = 0
            while offset < len(detail_rows):
                row = detail_rows[offset]
                line_count = len(self.wrap_pdf_text(row.get("itemName") or row.get("category2") or "-", 370, 7))
                row_height = max(44, 28 + line_count * 11)
                if page_rows and used_height + row_height > 620:
                    break
                page_rows.append((row, line_count, row_height))
                used_height += row_height
                offset += 1
            pages.append(self.pdf_detail_page(page_rows, offset - len(page_rows), len(detail_rows)))
        return self.make_pdf(pages)

    def build_pdf_legacy(self, rows):
        """検索結果を簡易PDFへ変換する。"""
        lines = ["レシート検索結果レポート", f"出力日時: {datetime.now().strftime('%Y/%m/%d %H:%M')}"]
        total, _, _ = self.summary(rows)
        lines.extend([f"明細件数: {len(rows)}", f"合計金額: {self.format_yen(total)}", ""])
        for row in rows:
            lines.append(
                f"{row['receiptDate']} {row['supplierName']} {row['itemName']} {self.format_yen(row['totalPrice'])}"
            )
        return self.simple_pdf(lines)

    def simple_pdf(self, lines):
        """標準フォントだけで読める最小構成のPDFを作成する。"""
        page_w, page_h = 595, 842
        parts = []
        y = page_h - 48
        for line in lines[:42]:
            parts.append(f"BT /F1 10 Tf 1 0 0 1 40 {y} Tm {self.pdf_hex(line)} Tj ET")
            y -= 18
        stream = "\n".join(parts).encode("utf-8")
        objects = [
            b"<< /Type /Catalog /Pages 2 0 R >>",
            b"<< /Type /Pages /Kids [6 0 R] /Count 1 >>",
            b"<< /Type /Font /Subtype /Type0 /BaseFont /HeiseiKakuGo-W5 /Encoding /UniJIS-UCS2-H /DescendantFonts [4 0 R] >>",
            b"<< /Type /Font /Subtype /CIDFontType0 /BaseFont /HeiseiKakuGo-W5 /CIDSystemInfo << /Registry (Adobe) /Ordering (Japan1) /Supplement 6 >> /DW 1000 >>",
            b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream",
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_w} {page_h}] /Resources << /Font << /F1 3 0 R >> >> /Contents 5 0 R >>".encode("ascii"),
        ]
        output = b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n"
        offsets = [0]
        for index, obj in enumerate(objects, start=1):
            offsets.append(len(output))
            output += f"{index} 0 obj\n".encode("ascii") + obj + b"\nendobj\n"
        xref = len(output)
        output += f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n".encode("ascii")
        for offset in offsets[1:]:
            output += f"{offset:010d} 00000 n \n".encode("ascii")
        output += f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n".encode("ascii")
        return output

    def pdf_cover_page(self, rows, total, category_totals, month_totals):
        commands = []
        self.pdf_rect(commands, 0, 0, 595, 842, fill="F8FAFC")
        self.pdf_rect(commands, 36, 732, 523, 74, fill="17324D")
        self.pdf_text(commands, 56, 778, "レシート検索結果レポート", 22, color="FFFFFF")
        self.pdf_text(commands, 56, 754, f"出力日時 {datetime.now().strftime('%Y/%m/%d %H:%M')}", 10, color="DDEAFE")

        x = 36
        for label, value in [
            ("明細件数", f"{len(rows):,} 件"),
            ("合計金額", self.format_yen(total)),
            ("分類数", f"{len(category_totals):,} 件"),
        ]:
            self.pdf_rect(commands, x, 662, 165, 52, fill="FFFFFF", stroke="D8E1EA")
            self.pdf_text(commands, x + 14, 692, label, 9, color="64748B")
            self.pdf_text(commands, x + 14, 674, value, 15, color="111827")
            x += 179

        self.pdf_section_title(commands, 36, 624, "分類別サマリー")
        y = 596
        for category, amount in sorted(category_totals.items(), key=lambda item: item[1], reverse=True)[:10]:
            self.pdf_text(commands, 48, y, self.truncate(category, 22), 10, color="111827")
            self.pdf_text(commands, 200, y, self.format_yen(amount), 10, color="111827")
            width = 150 * (amount / max(total, 1))
            self.pdf_rect(commands, 48, y - 13, 150, 5, fill="E2E8F0")
            self.pdf_rect(commands, 48, y - 13, max(width, 4), 5, fill="2563EB")
            y -= 27

        self.pdf_section_title(commands, 316, 624, "月別サマリー")
        y = 596
        for month, amount in sorted(month_totals.items())[-10:]:
            self.pdf_text(commands, 328, y, month, 10, color="111827")
            self.pdf_text(commands, 452, y, self.format_yen(amount), 10, color="111827")
            y -= 22

        return "\n".join(commands)

    def pdf_detail_page(self, rows, offset, total_count):
        commands = []
        self.pdf_rect(commands, 0, 0, 595, 842, fill="FFFFFF")
        self.pdf_text(commands, 36, 800, "明細一覧", 17, color="17324D")
        self.pdf_text(commands, 452, 800, f"{offset + 1}-{offset + len(rows)} / {total_count}", 9, color="64748B")
        columns = [
            ("日付", 48, 70),
            ("時間", 124, 38),
            ("店舗", 168, 132),
            ("種類", 306, 156),
            ("金額", 486, 62),
        ]
        self.pdf_rect(commands, 36, 764, 523, 24, fill="EFF6FF", stroke="D8E1EA")
        for header, x, _width in columns:
            self.pdf_text(commands, x, 772, header, 9, color="17324D")

        y = 742
        for index, (row, _line_count, row_height) in enumerate(rows):
            if index % 2 == 0:
                self.pdf_rect(commands, 36, y - row_height + 14, 523, row_height, fill="F8FAFC")
            item_lines = self.wrap_pdf_text(row["itemName"] or row["category2"] or "-", 370, 7)
            category = " / ".join([value for value in [row["category1"], row["category2"]] if value]) or "未分類"
            values = [
                row["receiptDate"] or "-",
                row["receiptTime"] or "-",
                row["supplierName"] or "インボイスなし",
                category,
                self.format_yen(row["totalPrice"]),
            ]
            for value, (_header, x, width) in zip(values, columns):
                self.pdf_text(commands, x, y, self.fit_pdf_text(value, width, 7), 7, color="111827")
            self.pdf_text(commands, 48, y - 15, "商品", 6, color="64748B")
            for line_index, line in enumerate(item_lines):
                self.pdf_text(commands, 80, y - 15 - line_index * 11, line, 7, color="111827")
            y -= row_height
        return "\n".join(commands)

    def pdf_section_title(self, commands, x, y, title):
        self.pdf_text(commands, x, y, title, 13, color="17324D")
        self.pdf_line(commands, x, y - 8, x + 220, y - 8, color="D8E1EA")

    def pdf_text(self, commands, x, y, text, size=10, color="111827"):
        r, g, b = self.pdf_color(color)
        cursor = x
        for value, is_ascii in self.pdf_text_runs(text):
            font = "FL" if is_ascii else "FJ"
            encoded = self.pdf_literal(value) if is_ascii else self.pdf_hex(value)
            commands.append(
                f"BT /{font} {size} Tf {r:.3f} {g:.3f} {b:.3f} rg "
                f"1 0 0 1 {cursor:.2f} {y} Tm {encoded} Tj ET"
            )
            cursor += self.pdf_text_width(value, size)

    def pdf_rect(self, commands, x, y, width, height, fill=None, stroke=None):
        if fill:
            r, g, b = self.pdf_color(fill)
            commands.append(f"{r:.3f} {g:.3f} {b:.3f} rg {x} {y} {width} {height} re f")
        if stroke:
            r, g, b = self.pdf_color(stroke)
            commands.append(f"{r:.3f} {g:.3f} {b:.3f} RG {x} {y} {width} {height} re S")

    def pdf_line(self, commands, x1, y1, x2, y2, color="D8E1EA"):
        r, g, b = self.pdf_color(color)
        commands.append(f"{r:.3f} {g:.3f} {b:.3f} RG {x1} {y1} m {x2} {y2} l S")

    def pdf_color(self, hex_color):
        value = hex_color.lstrip("#")
        return tuple(int(value[index:index + 2], 16) / 255 for index in (0, 2, 4))

    def make_pdf(self, page_streams):
        page_w, page_h = 595, 842
        objects = [
            b"<< /Type /Catalog /Pages 2 0 R >>",
            b"",
            b"<< /Type /Font /Subtype /Type0 /BaseFont /HeiseiKakuGo-W5 /Encoding /UniJIS-UCS2-H /DescendantFonts [4 0 R] >>",
            b"<< /Type /Font /Subtype /CIDFontType0 /BaseFont /HeiseiKakuGo-W5 "
            b"/CIDSystemInfo << /Registry (Adobe) /Ordering (Japan1) /Supplement 6 >> "
            b"/FontDescriptor 6 0 R /DW 1000 >>",
            b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
            b"<< /Type /FontDescriptor /FontName /HeiseiKakuGo-W5 /Flags 4 "
            b"/FontBBox [-92 -250 1010 922] /ItalicAngle 0 /Ascent 880 "
            b"/Descent -120 /CapHeight 737 /StemV 80 >>",
        ]
        page_refs = []
        for page in page_streams:
            stream = page.encode("utf-8")
            content_ref = len(objects) + 1
            objects.append(b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream")
            page_ref = len(objects) + 1
            page_refs.append(page_ref)
            objects.append(
                f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_w} {page_h}] "
                f"/Resources << /Font << /FJ 3 0 R /FL 5 0 R >> >> /Contents {content_ref} 0 R >>".encode("ascii")
            )
        kids = " ".join(f"{ref} 0 R" for ref in page_refs)
        objects[1] = f"<< /Type /Pages /Kids [{kids}] /Count {len(page_refs)} >>".encode("ascii")

        output = b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n"
        offsets = [0]
        for index, obj in enumerate(objects, start=1):
            offsets.append(len(output))
            output += f"{index} 0 obj\n".encode("ascii") + obj + b"\nendobj\n"
        xref = len(output)
        output += f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n".encode("ascii")
        for offset in offsets[1:]:
            output += f"{offset:010d} 00000 n \n".encode("ascii")
        output += f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n".encode("ascii")
        return output

    def truncate(self, value, limit):
        text = self.text(value)
        return text if len(text) <= limit else text[:limit - 1] + "…"

    def wrap_pdf_text(self, value, max_width, size=10):
        text = self.text(value)
        if not text:
            return ["-"]
        lines = []
        current = ""
        for char in text:
            candidate = current + char
            if current and self.pdf_text_width(candidate, size) > max_width:
                lines.append(current)
                current = char
            else:
                current = candidate
        if current:
            lines.append(current)
        return lines or ["-"]

    def fit_pdf_text(self, value, max_width, size=10):
        text = self.text(value)
        if not text or self.pdf_text_width(text, size) <= max_width:
            return text
        marker = "..."
        fitted = ""
        for char in text:
            if self.pdf_text_width(fitted + char + marker, size) > max_width:
                break
            fitted += char
        return fitted + marker if fitted else marker

    def pdf_text_width(self, value, size=10):
        width = 0
        for char in self.text(value):
            width += size * (0.55 if ord(char) < 128 else 1.0)
        return width

    def pdf_text_runs(self, value):
        """Split mixed Japanese/ASCII text so each script uses a suitable font."""
        text = self.text(value)
        if not text:
            return []
        runs = []
        current = text[0]
        current_ascii = ord(text[0]) < 128
        for char in text[1:]:
            is_ascii = ord(char) < 128
            if is_ascii == current_ascii:
                current += char
            else:
                runs.append((current, current_ascii))
                current = char
                current_ascii = is_ascii
        runs.append((current, current_ascii))
        return runs

    def summary(self, rows):
        """出力用データから総額、分類別、月別の集計を作る。"""
        total = sum(self.number(row.get("totalPrice")) for row in rows)
        category_totals = {}
        month_totals = {}
        for row in rows:
            category = row.get("category1") or "未分類"
            category_totals[category] = category_totals.get(category, 0) + self.number(row.get("totalPrice"))
            month = (row.get("receiptDate") or "不明")[:7] or "不明"
            month_totals[month] = month_totals.get(month, 0) + self.number(row.get("totalPrice"))
        return total, category_totals, month_totals

    def pdf_hex(self, text):
        """PDFの日本語文字列として埋め込める16進文字列へ変換する。"""
        return "<" + self.text(text).encode("utf-16-be").hex().upper() + ">"

    def pdf_literal(self, text):
        """Escape an ASCII string for a PDF literal string."""
        value = self.text(text).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        return f"({value})"

    def text(self, value):
        """Noneを空文字に寄せて出力用文字列へ変換する。"""
        return "" if value is None else str(value).strip()

    def number(self, value):
        """出力集計で扱う数値を安全にfloatへ変換する。"""
        try:
            if value is None or value == "":
                return 0
            return float(value)
        except (TypeError, ValueError):
            return 0

    def format_yen(self, value):
        """円表記用に小数を丸めてカンマ付き文字列へ変換する。"""
        return f"{int(round(self.number(value))):,}円"

    def format_tax_rate(self, value):
        """税率を百分率表記へ変換する。"""
        rate = self.number(value)
        if rate <= 0:
            return ""
        return f"{round(rate * 100):g}%"
