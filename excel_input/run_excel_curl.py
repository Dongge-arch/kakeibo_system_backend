import argparse
import json
import os
import subprocess
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_REQUEST_HEADERS = {
    "Content-Type": "application/json",
    "Accept": "application/json",
    "x-kakeibo-user-id": "",
    "x-kakeibo-user-email": "",
    "x-kakeibo-user-name": "",
    "x-kakeibo-user-nickname": "",
    "Authorization": "",
}


def load_default_headers():
    headers = dict(DEFAULT_REQUEST_HEADERS)
    env_map = {
        "x-kakeibo-user-id": "KAKEIBO_CURL_USER_ID",
        "x-kakeibo-user-email": "KAKEIBO_CURL_USER_EMAIL",
        "x-kakeibo-user-name": "KAKEIBO_CURL_USER_NAME",
        "x-kakeibo-user-nickname": "KAKEIBO_CURL_USER_NICKNAME",
        "Authorization": "KAKEIBO_CURL_AUTHORIZATION",
    }
    for header, env_name in env_map.items():
        value = os.getenv(env_name, "").strip()
        if value:
            headers[header] = value
    return headers


def parse_extra_header(raw_header):
    if ":" not in raw_header:
        raise argparse.ArgumentTypeError("extra header must be in KEY: VALUE format")
    key, value = raw_header.split(":", 1)
    key = key.strip()
    value = value.strip()
    if not key:
        raise argparse.ArgumentTypeError("extra header key cannot be empty")
    return key, value


def parse_args():
    parser = argparse.ArgumentParser(
        description="Read receipt sheets from Excel and execute curl POSTs for each row."
    )
    parser.add_argument("--excel", required=True, help="Path to the Excel workbook.")
    parser.add_argument(
        "--target-url",
        default="http://localhost:8000/receipt/newReceiptRegistration",
        help="Target API endpoint for the curl POST.",
    )
    parser.add_argument("--api-key", default=None, help="Optional x-api-key header value.")
    parser.add_argument("--user-id", default=None, help="Override x-kakeibo-user-id header.")
    parser.add_argument("--user-email", default=None, help="Override x-kakeibo-user-email header.")
    parser.add_argument("--user-name", default=None, help="Override x-kakeibo-user-name header.")
    parser.add_argument(
        "--user-nickname",
        default=None,
        help="Override x-kakeibo-user-nickname header.",
    )
    parser.add_argument(
        "--authorization",
        default=None,
        help="Override Authorization header, e.g. Bearer <token>.",
    )
    parser.add_argument(
        "--extra-header",
        action="append",
        default=None,
        metavar="KEY: VALUE",
        help="Additional custom header. Repeat to add multiple headers.",
    )
    parser.add_argument("--log-file", default=None, help="Optional log file path.")
    return parser.parse_args()


def build_request_headers(args):
    headers = load_default_headers()

    if args.user_id:
        headers["x-kakeibo-user-id"] = args.user_id.strip()
    if args.user_email:
        headers["x-kakeibo-user-email"] = args.user_email.strip()
    if args.user_name:
        headers["x-kakeibo-user-name"] = args.user_name.strip()
    if args.user_nickname:
        headers["x-kakeibo-user-nickname"] = args.user_nickname.strip()
    if args.authorization:
        headers["Authorization"] = args.authorization.strip()

    if args.extra_header:
        for raw_header in args.extra_header:
            key, value = parse_extra_header(raw_header)
            headers[key] = value

    if args.api_key:
        headers["x-api-key"] = args.api_key.strip()

    return {key: value for key, value in headers.items() if value}


def build_command(target_url, body, headers=None):
    command = ["curl", "-s", "-X", "POST"]
    request_headers = dict(headers or {})
    for key, value in request_headers.items():
        command.extend(["-H", f"{key}: {value}"])
    command.extend(["-d", body, target_url])
    return command


def normalize_scalar(value):
    if isinstance(value, (datetime, date, time)):
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        return value.strftime("%H:%M:%S")
    if isinstance(value, (int, float)):
        return value
    if value is None:
        return None
    return str(value).strip()


def convert_datetime(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return None


def convert_time(value):
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return None


def convert_number(value):
    if value in (None, ""):
        return None
    try:
        number = float(str(value).replace(",", "").strip())
        if number.is_integer():
            return int(number)
        return number
    except Exception:
        return None


def find_indices(row, target):
    positions = []
    for index, cell in enumerate(row):
        if isinstance(cell, str) and cell.strip() == target:
            positions.append(index)
    return positions


def build_receipt_payload(sheet, row_values, header_row):
    header = [cell.value for cell in header_row]
    index_map = {value: idx for idx, value in enumerate(header)}

    invoice_idx = index_map.get("登録者番号")
    if invoice_idx is None:
        return None

    invoice_number = normalize_scalar(row_values[invoice_idx])
    if not invoice_number:
        return None

    right_date_idx = None
    for candidate in range(invoice_idx - 1, -1, -1):
        candidate_value = row_values[candidate]
        if isinstance(candidate_value, (datetime, date)):
            right_date_idx = candidate
            break
    if right_date_idx is None:
        right_date_idx = invoice_idx - 1

    place_idx = index_map.get("場所")
    item_idx = None
    if place_idx is not None:
        for candidate in range(place_idx + 1, len(row_values)):
            if isinstance(header[candidate], str) and header[candidate].strip() == "内容":
                item_idx = candidate
                break
    if item_idx is None:
        item_idx = index_map.get("内容")

    category_idx = None
    for candidate in range(invoice_idx + 1, len(row_values)):
        if isinstance(header[candidate], str) and header[candidate].strip() == "分類":
            category_idx = candidate
            break
    if category_idx is None:
        category_idx = invoice_idx + 1

    tax_flag_idx = index_map.get("軽減税率")
    quantity_idx = index_map.get("点数")
    unit_price_idx = index_map.get("単価（税込）")
    total_price_idx = index_map.get("金額（税込）")

    purchase_datetime = row_values[right_date_idx]
    receipt_date = convert_datetime(purchase_datetime)
    receipt_time = convert_time(purchase_datetime)

    supplier_name = normalize_scalar(row_values[place_idx]) if place_idx is not None else ""
    item_name = normalize_scalar(row_values[item_idx]) if item_idx is not None else ""
    category1 = normalize_scalar(row_values[category_idx]) if category_idx is not None else ""
    tax_flag_value = normalize_scalar(row_values[tax_flag_idx]) if tax_flag_idx is not None else ""
    quantity = convert_number(row_values[quantity_idx]) if quantity_idx is not None else None
    unit_price = convert_number(row_values[unit_price_idx]) if unit_price_idx is not None else None
    total_price = convert_number(row_values[total_price_idx]) if total_price_idx is not None else None

    detail = {
        "itemName": item_name or "不明商品",
        "category1": category1 or "",
        "category2": "",
        "quantity": quantity if quantity is not None else 1,
        "unitPrice": unit_price if unit_price is not None else total_price if total_price is not None else 0,
        "totalPrice": total_price if total_price is not None else unit_price if unit_price is not None else 0,
    }

    receipt_info = {
        "invoiceRegistrationNumber": invoice_number,
        "supplierName": supplier_name,
        "receiptDate": receipt_date,
        "receiptTime": receipt_time,
        "taxFlag": 1 if str(tax_flag_value) not in ("", "0", "0.0", "False", "false") else 0,
        "receiptDetailCount": 1,
        "receiptDetails": [detail],
        "totalPrice": total_price if total_price is not None else detail["totalPrice"],
    }

    return {"receiptInfo": receipt_info}


def main():
    args = parse_args()
    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    workbook = load_workbook(excel_path, data_only=True)
    sheet_names = [sheet.title.strip() for sheet in workbook.worksheets]
    month_sheets = [
        sheet for sheet in workbook.worksheets
        if sheet.title.strip() not in {
            "初期設定",
            "使い方(サンプル)",
            "場所リスト",
            "支援金口座引き出し記録表",
            "3月収支図示",
            "推移要約",
            "推移詳細",
        }
        and not sheet.title.strip().startswith("グラフ")
        and sheet.title.strip() in {"2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月"}
    ]

    log_path = Path(args.log_file) if args.log_file else Path(__file__).with_suffix(".log")
    log_path.parent.mkdir(parents=True, exist_ok=True)

    with log_path.open("a", encoding="utf-8") as log_handle:
        for sheet in month_sheets:
            header_row = None
            header_row_index = None
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                values = [cell.value for cell in row]
                if any(isinstance(value, str) and value.strip() == "登録者番号" for value in values):
                    header_row = row
                    header_row_index = row[0].row
                    break
            if header_row is None or header_row_index is None:
                continue

            for row_num, row in enumerate(sheet.iter_rows(min_row=header_row_index + 1, max_row=sheet.max_row), start=header_row_index + 1):
                row_values = [cell.value for cell in row]
                if not any(value not in (None, "") for value in row_values):
                    continue

                payload = build_receipt_payload(sheet, row_values, header_row)
                if not payload:
                    continue

                body = json.dumps(payload, ensure_ascii=False, default=normalize_scalar)
                request_headers = build_request_headers(args)
                command = build_command(args.target_url, body, request_headers)

                print(f"Executing sheet={sheet.title} row={row_num}: {' '.join(command)}")
                result = subprocess.run(command, capture_output=True, text=True)
                status = result.returncode
                response_text = (result.stdout or result.stderr or "").strip()
                log_handle.write(
                    json.dumps(
                        {
                            "sheet": sheet.title,
                            "row": row_num,
                            "command": command,
                            "returncode": status,
                            "response": response_text,
                        },
                        ensure_ascii=False,
                        default=normalize_scalar,
                    )
                    + "\n"
                )
                log_handle.flush()

                if status != 0:
                    print(f"curl failed for sheet={sheet.title} row={row_num}, returncode={status}")
                    print(response_text)
                else:
                    print(f"curl success for sheet={sheet.title} row={row_num}")


if __name__ == "__main__":
    main()
