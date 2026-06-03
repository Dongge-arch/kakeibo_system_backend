from openpyxl import load_workbook

wb = load_workbook('C:/Users/董 昊哲/Desktop/kakeibo_2026.xlsx', data_only=True)
month_sheets = []
for sheet in wb.worksheets:
    name = sheet.title.strip()
    if name in {
        '初期設定', '使い方(サンプル)', '場所リスト', '支援金口座引き出し記録表',
        '3月収支図示', '推移要約', '推移詳細'
    }:
        continue
    if name.startswith('グラフ'):
        continue
    if name in {'2月','3月','4月','5月','6月','7月','8月','9月','10月','11月','12月'}:
        month_sheets.append(sheet)

print('month sheets', [s.title for s in month_sheets])
for sheet in month_sheets[:3]:
    print('SHEET', sheet.title)
    for row in sheet.iter_rows(min_row=4, max_row=min(10, sheet.max_row), values_only=True):
        print(row)
    print('----')
